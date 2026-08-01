from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from src.presentation.posting_workbook import (
    PackageLine,
    PackageSummary,
    build_workbook,
)


def summary() -> PackageSummary:
    return PackageSummary(
        package_id="pkg1",
        payout_id="po1",
        processor="Stripe",
        processor_account="Main Guesty",
        processor_payout_date="2026-07-08",
        bank_transaction_id="bank1",
        bank_transaction_date="2026-07-08",
        bank_description="STRIPE PAYOUT",
        bank_amount=134.32,
        payout_amount=134.32,
        posting_total=134.32,
        difference=0.0,
        bank_difference=0.0,
        balanced="Yes",
        bank_balanced="Yes",
        review_status="Ready for Review",
        confidence="Ready",
        comparison_status="Improved",
        posting_line_count=2,
        payment_event_count=2,
        source_count=2,
        reversal_line_count=1,
        seed_line_count=0,
        sheet_name="Stripe - 2026-07-08",
        bank_feed_label="2026-07-08 | STRIPE PAYOUT | 134.32",
        review_notes="Includes one reversal.",
    )


def lines() -> list[PackageLine]:
    return [
        PackageLine(
            package_id="pkg1",
            payout_id="po1",
            line_number=1,
            account="Revenue",
            qb_class="Hospitality",
            description="Revenue",
            amount=188.64,
            posting_type="Original",
            ledger_source="Persistent History",
            line_note="Persistent accounting history.",
        ),
        PackageLine(
            package_id="pkg1",
            payout_id="po1",
            line_number=2,
            account="RV Rent",
            qb_class="RV Sites",
            description="Refund",
            amount=-54.32,
            posting_type="Reversal",
            ledger_source="Reversal Preview",
            line_note="Historical reversal.",
        ),
    ]


def _find_cell_by_value(sheet, value: str):
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == value:
                return cell
    return None


def test_navigation_uses_internal_locations(
    tmp_path: Path,
):
    workbook = build_workbook([summary()], lines())
    path = tmp_path / "package.xlsx"
    workbook.save(path)

    reopened = load_workbook(
        path,
        data_only=False,
    )

    dashboard_link = reopened[
        "Dashboard"
    ]["A10"].hyperlink

    payout_sheet = reopened[
        "Stripe - 2026-07-08"
    ]
    back_cell = _find_cell_by_value(
        payout_sheet,
        "Back to Dashboard",
    )

    assert dashboard_link is not None
    assert dashboard_link.location == (
        "'Stripe - 2026-07-08'!A1"
    )
    assert dashboard_link.target is None

    assert back_cell is not None
    assert back_cell.hyperlink is not None
    assert back_cell.hyperlink.location == (
        "'Dashboard'!A1"
    )
    assert back_cell.hyperlink.target is None
