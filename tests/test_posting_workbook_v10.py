from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from src.presentation.posting_workbook import (
    PackageLine,
    PackageSummary,
    build_workbook,
    output_filename,
    package_month_label,
)


def summary(
    bank_date: str,
) -> PackageSummary:
    return PackageSummary(
        package_id="pkg1",
        payout_id="po1",
        processor="Stripe",
        processor_account="Main Guesty",
        processor_payout_date="2026-07-08",
        bank_transaction_id="bank1",
        bank_transaction_date=bank_date,
        bank_description="STRIPE PAYOUT",
        bank_amount=134.32,
        payout_amount=134.32,
        posting_total=134.32,
        difference=0.0,
        bank_difference=0.0,
        balanced="Yes",
        bank_balanced="Yes",
        review_status="Ready for Review",
        confidence="Ready",
        comparison_status="Improved",
        posting_line_count=3,
        payment_event_count=4,
        source_count=3,
        reversal_line_count=2,
        seed_line_count=0,
        sheet_name="Stripe - 2026-07-08",
        bank_feed_label=(
            "2026-07-08 | STRIPE PAYOUT | 134.32"
        ),
        review_notes="Includes reversal lines.",
    )


def lines() -> list[PackageLine]:
    return [
        PackageLine(
            package_id="pkg1",
            payout_id="po1",
            line_number=1,
            account="Motel Rent - Short Term",
            qb_class="Hospitality",
            description="Room revenue",
            amount=188.64,
            posting_type="Original",
            ledger_source="Persistent History",
            line_note="Persistent accounting history.",
        ),
        PackageLine(
            package_id="pkg1",
            payout_id="po1",
            line_number=2,
            account="RV Rent - Nightly",
            qb_class="RV Sites",
            description="Refund",
            amount=-54.32,
            posting_type="Reversal",
            ledger_source="Reversal Preview",
            line_note=(
                "Historical reversal. Generated from "
                "posting-history reversal."
            ),
        ),
    ]


def test_month_label_uses_latest_bank_date():
    summaries = [
        summary("2026-06-29"),
        summary("2026-07-08"),
    ]
    assert package_month_label(
        summaries
    ) == "2026-07"


def test_output_filename_is_month_specific():
    summaries = [summary("2026-07-08")]
    assert output_filename(
        summaries
    ) == (
        "QuickBooks_Posting_Package_"
        "2026-07.xlsx"
    )


def test_workbook_contains_dashboard_and_payout_sheet(
    tmp_path: Path,
):
    workbook = build_workbook(
        [summary("2026-07-08")],
        lines(),
    )
    path = tmp_path / "package.xlsx"
    workbook.save(path)

    reopened = load_workbook(
        path,
        data_only=False,
    )

    assert reopened.sheetnames == [
        "Dashboard",
        "Stripe - 2026-07-08",
    ]

    dashboard = reopened["Dashboard"]
    payout = reopened["Stripe - 2026-07-08"]

    assert dashboard["A1"].value.startswith(
        "Dark Sky River Lodge"
    )
    assert dashboard["A10"].value == "Open"
    assert dashboard["A10"].hyperlink is not None

    assert payout["A1"].value == (
        "Stripe Bank-Feed Split Instructions"
    )
    assert payout["B5"].value == 134.32
    assert payout["E13"].value == 188.64
    assert payout["E14"].value == -54.32
