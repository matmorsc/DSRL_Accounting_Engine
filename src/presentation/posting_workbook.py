from __future__ import annotations

import csv
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.table import Table, TableStyleInfo


TITLE_FILL = "17324D"
SECTION_FILL = "315F7D"
LIGHT_FILL = "E5EEF5"
NOTE_FILL = "F8FAFC"
READY_FILL = "DCFCE7"
READY_FONT = "166534"
CAUTION_FILL = "FEF3C7"
CAUTION_FONT = "92400E"
REVIEW_FILL = "FEE2E2"
REVIEW_FONT = "991B1B"
WHITE = "FFFFFF"

CURRENCY_FORMAT = '$#,##0.00;[Red]-$#,##0.00'
DATE_FORMAT = "yyyy-mm-dd"


@dataclass(frozen=True)
class PackageSummary:
    package_id: str
    payout_id: str
    processor: str
    processor_account: str
    processor_payout_date: str
    bank_transaction_id: str
    bank_transaction_date: str
    bank_description: str
    bank_amount: float
    payout_amount: float
    posting_total: float
    difference: float
    bank_difference: float
    balanced: str
    bank_balanced: str
    review_status: str
    confidence: str
    comparison_status: str
    posting_line_count: int
    payment_event_count: int
    source_count: int
    reversal_line_count: int
    seed_line_count: int
    sheet_name: str
    bank_feed_label: str
    review_notes: str


@dataclass(frozen=True)
class PackageLine:
    package_id: str
    payout_id: str
    line_number: int
    account: str
    qb_class: str
    description: str
    amount: float
    posting_type: str
    ledger_source: str
    line_note: str


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _float(value: object) -> float:
    try:
        return round(float(value), 2)
    except (TypeError, ValueError):
        return 0.0


def _int(value: object) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def load_summary(path: Path) -> list[PackageSummary]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = []
        for row in csv.DictReader(handle):
            rows.append(
                PackageSummary(
                    package_id=_text(row.get("package_id")),
                    payout_id=_text(row.get("payout_id")),
                    processor=_text(row.get("processor")),
                    processor_account=_text(row.get("processor_account")),
                    processor_payout_date=_text(
                        row.get("processor_payout_date")
                    ),
                    bank_transaction_id=_text(
                        row.get("bank_transaction_id")
                    ),
                    bank_transaction_date=_text(
                        row.get("bank_transaction_date")
                    ),
                    bank_description=_text(
                        row.get("bank_description")
                    ),
                    bank_amount=_float(row.get("bank_amount")),
                    payout_amount=_float(row.get("payout_amount")),
                    posting_total=_float(row.get("posting_total")),
                    difference=_float(row.get("difference")),
                    bank_difference=_float(
                        row.get("bank_difference")
                    ),
                    balanced=_text(row.get("balanced")),
                    bank_balanced=_text(row.get("bank_balanced")),
                    review_status=_text(row.get("review_status")),
                    confidence=_text(row.get("confidence")),
                    comparison_status=_text(
                        row.get("comparison_status")
                    ),
                    posting_line_count=_int(
                        row.get("posting_line_count")
                    ),
                    payment_event_count=_int(
                        row.get("payment_event_count")
                    ),
                    source_count=_int(row.get("source_count")),
                    reversal_line_count=_int(
                        row.get("reversal_line_count")
                    ),
                    seed_line_count=_int(
                        row.get("seed_line_count")
                    ),
                    sheet_name=_text(row.get("sheet_name")),
                    bank_feed_label=_text(
                        row.get("bank_feed_label")
                    ),
                    review_notes=_text(row.get("review_notes")),
                )
            )
        return rows


def load_lines(path: Path) -> list[PackageLine]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = []
        for row in csv.DictReader(handle):
            rows.append(
                PackageLine(
                    package_id=_text(row.get("package_id")),
                    payout_id=_text(row.get("payout_id")),
                    line_number=_int(row.get("line_number")),
                    account=_text(row.get("account")),
                    qb_class=_text(row.get("class")),
                    description=_text(row.get("description")),
                    amount=_float(row.get("amount")),
                    posting_type=_text(row.get("posting_type")),
                    ledger_source=_text(row.get("ledger_source")),
                    line_note=_text(row.get("line_note")),
                )
            )
        return rows


def package_month_label(
    summaries: Iterable[PackageSummary],
) -> str:
    dates = sorted(
        summary.bank_transaction_date
        for summary in summaries
        if summary.bank_transaction_date
    )
    if not dates:
        return "Undated"

    parsed = datetime.strptime(dates[-1], "%Y-%m-%d")
    return parsed.strftime("%Y-%m")


def output_filename(
    summaries: Iterable[PackageSummary],
) -> str:
    return (
        f"QuickBooks_Posting_Package_"
        f"{package_month_label(summaries)}.xlsx"
    )


def _status_colors(confidence: str) -> tuple[str, str]:
    if confidence == "Ready":
        return READY_FILL, READY_FONT
    if confidence.startswith("Ready"):
        return CAUTION_FILL, CAUTION_FONT
    return REVIEW_FILL, REVIEW_FONT


def _style_header(cells) -> None:
    for cell in cells:
        cell.fill = PatternFill("solid", fgColor=TITLE_FILL)
        cell.font = Font(bold=True, color=WHITE)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )


def _style_section_header(cell) -> None:
    cell.fill = PatternFill("solid", fgColor=SECTION_FILL)
    cell.font = Font(bold=True, color=WHITE)


def _set_widths(sheet, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _set_internal_hyperlink(
    cell,
    *,
    location: str,
    display: str,
) -> None:
    cell.value = display
    cell.hyperlink = Hyperlink(
        ref=cell.coordinate,
        location=location,
        display=display,
    )
    cell.style = "Hyperlink"


def _write_dashboard(
    workbook: Workbook,
    summaries: list[PackageSummary],
) -> None:
    sheet = workbook.create_sheet("Dashboard")

    ready = sum(
        1 for row in summaries if row.confidence == "Ready"
    )
    needs_review = len(summaries) - ready
    total_bank_amount = round(
        sum(row.bank_amount for row in summaries),
        2,
    )
    total_posting = round(
        sum(row.posting_total for row in summaries),
        2,
    )

    sheet.merge_cells("A1:J1")
    sheet["A1"] = (
        "Dark Sky River Lodge — QuickBooks Posting Package"
    )
    sheet["A1"].fill = PatternFill(
        "solid",
        fgColor=TITLE_FILL,
    )
    sheet["A1"].font = Font(
        bold=True,
        color=WHITE,
        size=18,
    )
    sheet["A1"].alignment = Alignment(
        horizontal="left",
        vertical="center",
    )
    sheet.row_dimensions[1].height = 30

    metrics_left = [
        ("Metric", "Value"),
        ("Payout packages", len(summaries)),
        ("Ready", ready),
        ("Needs Review", needs_review),
    ]
    metrics_right = [
        ("Metric", "Value"),
        ("Bank amount", total_bank_amount),
        ("Posting total", total_posting),
        (
            "Difference",
            round(total_posting - total_bank_amount, 2),
        ),
    ]

    for row_index, values in enumerate(
        metrics_left,
        start=3,
    ):
        sheet.cell(row=row_index, column=1, value=values[0])
        sheet.cell(row=row_index, column=2, value=values[1])

    for row_index, values in enumerate(
        metrics_right,
        start=3,
    ):
        sheet.cell(row=row_index, column=4, value=values[0])
        sheet.cell(row=row_index, column=5, value=values[1])

    _style_header(sheet[3][0:2])
    _style_header(sheet[3][3:5])

    for row in range(4, 7):
        sheet.cell(row=row, column=5).number_format = (
            CURRENCY_FORMAT
        )

    headers = [
        "Open",
        "Status",
        "Processor",
        "Bank Date",
        "Bank Description",
        "Bank Amount",
        "Posting Total",
        "Difference",
        "Payout ID",
        "Notes",
    ]
    start_row = 9

    for column, value in enumerate(headers, start=1):
        sheet.cell(
            row=start_row,
            column=column,
            value=value,
        )
    _style_header(sheet[start_row])

    first_data_row = start_row + 1

    for row_index, summary in enumerate(
        summaries,
        start=first_data_row,
    ):
        _set_internal_hyperlink(
            sheet.cell(
                row=row_index,
                column=1,
            ),
            location=f"'{summary.sheet_name}'!A1",
            display="Open",
        )

        values = [
            summary.confidence,
            summary.processor,
            summary.bank_transaction_date,
            summary.bank_description,
            summary.bank_amount,
            summary.posting_total,
            summary.bank_difference,
            summary.payout_id,
            summary.review_notes,
        ]

        for column, value in enumerate(
            values,
            start=2,
        ):
            sheet.cell(
                row=row_index,
                column=column,
                value=value,
            )

        fill, font_color = _status_colors(
            summary.confidence
        )
        status_cell = sheet.cell(
            row=row_index,
            column=2,
        )
        status_cell.fill = PatternFill(
            "solid",
            fgColor=fill,
        )
        status_cell.font = Font(
            bold=True,
            color=font_color,
        )
        status_cell.alignment = Alignment(
            horizontal="center",
        )

        for column in (6, 7, 8):
            sheet.cell(
                row=row_index,
                column=column,
            ).number_format = CURRENCY_FORMAT

        sheet.cell(
            row=row_index,
            column=4,
        ).number_format = DATE_FORMAT

        sheet.cell(
            row=row_index,
            column=5,
        ).alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )
        sheet.cell(
            row=row_index,
            column=10,
        ).alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )

    last_data_row = start_row + len(summaries)

    if summaries:
        table = Table(
            displayName="PostingPackageDashboard",
            ref=f"A{start_row}:J{last_data_row}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

        red_fill = PatternFill(
            "solid",
            fgColor=REVIEW_FILL,
        )
        red_font = Font(color=REVIEW_FONT)
        sheet.conditional_formatting.add(
            f"H{first_data_row}:H{last_data_row}",
            FormulaRule(
                formula=[
                    f"ABS(H{first_data_row})>0.02"
                ],
                fill=red_fill,
                font=red_font,
            ),
        )

    _set_widths(
        sheet,
        {
            "A": 9,
            "B": 18,
            "C": 12,
            "D": 13,
            "E": 42,
            "F": 15,
            "G": 15,
            "H": 15,
            "I": 32,
            "J": 45,
        },
    )

    sheet.freeze_panes = f"A{first_data_row}"
    sheet.auto_filter.ref = (
        f"A{start_row}:J{last_data_row}"
        if summaries
        else f"A{start_row}:J{start_row}"
    )


def _write_package_sheet(
    workbook: Workbook,
    summary: PackageSummary,
    lines: list[PackageLine],
) -> None:
    sheet = workbook.create_sheet(
        summary.sheet_name
    )

    sheet.merge_cells("A1:F1")
    sheet["A1"] = (
        f"{summary.processor} Bank-Feed Split Instructions"
    )
    sheet["A1"].fill = PatternFill(
        "solid",
        fgColor=TITLE_FILL,
    )
    sheet["A1"].font = Font(
        bold=True,
        color=WHITE,
        size=16,
    )
    sheet["A1"].alignment = Alignment(
        vertical="center",
    )
    sheet.row_dimensions[1].height = 28

    left_values = [
        ("Bank date", summary.bank_transaction_date),
        ("Bank description", summary.bank_description),
        ("Bank amount", summary.bank_amount),
        ("Posting total", summary.posting_total),
        ("Difference", summary.bank_difference),
        ("Confidence", summary.confidence),
        (
            "Processor payout date",
            summary.processor_payout_date,
        ),
        ("Payout ID", summary.payout_id),
    ]
    right_values = [
        ("Processor", summary.processor),
        (
            "Processor account",
            summary.processor_account,
        ),
        (
            "Bank transaction ID",
            summary.bank_transaction_id,
        ),
        ("Comparison", summary.comparison_status),
        ("Payment events", summary.payment_event_count),
        ("Sources", summary.source_count),
    ]

    for row_index, (label, value) in enumerate(
        left_values,
        start=3,
    ):
        sheet.cell(
            row=row_index,
            column=1,
            value=label,
        )
        sheet.cell(
            row=row_index,
            column=2,
            value=value,
        )

    for row_index, (label, value) in enumerate(
        right_values,
        start=3,
    ):
        sheet.cell(
            row=row_index,
            column=4,
            value=label,
        )
        sheet.cell(
            row=row_index,
            column=5,
            value=value,
        )

    _style_section_header(sheet["A3"])
    _style_section_header(sheet["D3"])

    for row in (5, 6, 7):
        sheet.cell(
            row=row,
            column=2,
        ).number_format = CURRENCY_FORMAT

    for row in range(3, 11):
        sheet.cell(
            row=row,
            column=2,
        ).alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )

    for row in range(3, 9):
        sheet.cell(
            row=row,
            column=5,
        ).alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )

    headers = [
        "Line",
        "QuickBooks Category",
        "Class",
        "Description",
        "Amount",
        "Why this line exists",
    ]

    for column, value in enumerate(
        headers,
        start=1,
    ):
        sheet.cell(
            row=12,
            column=column,
            value=value,
        )
    _style_header(sheet[12])

    ordered_lines = sorted(
        lines,
        key=lambda item: item.line_number,
    )
    first_line_row = 13

    for row_index, line in enumerate(
        ordered_lines,
        start=first_line_row,
    ):
        values = [
            line.line_number,
            line.account,
            line.qb_class,
            line.description,
            line.amount,
            line.line_note,
        ]

        for column, value in enumerate(
            values,
            start=1,
        ):
            sheet.cell(
                row=row_index,
                column=column,
                value=value,
            )

        sheet.cell(
            row=row_index,
            column=5,
        ).number_format = CURRENCY_FORMAT

        for column in (4, 6):
            sheet.cell(
                row=row_index,
                column=column,
            ).alignment = Alignment(
                wrap_text=True,
                vertical="top",
            )

    last_line_row = (
        first_line_row + len(ordered_lines) - 1
    )
    total_row = max(
        first_line_row,
        last_line_row + 2,
    )

    sheet.cell(
        row=total_row,
        column=4,
        value="Split total",
    )
    sheet.cell(
        row=total_row + 1,
        column=4,
        value="Difference to bank",
    )

    if ordered_lines:
        sheet.cell(
            row=total_row,
            column=5,
            value=(
                f"=SUM(E{first_line_row}:"
                f"E{last_line_row})"
            ),
        )
    else:
        sheet.cell(
            row=total_row,
            column=5,
            value=0.0,
        )

    sheet.cell(
        row=total_row + 1,
        column=5,
        value=f"=E{total_row}-$B$5",
    )

    for row in (total_row, total_row + 1):
        for column in (4, 5):
            cell = sheet.cell(
                row=row,
                column=column,
            )
            cell.fill = PatternFill(
                "solid",
                fgColor=LIGHT_FILL,
            )
            cell.font = Font(bold=True)
        sheet.cell(
            row=row,
            column=5,
        ).number_format = CURRENCY_FORMAT

    notes_row = total_row + 4
    sheet.merge_cells(
        start_row=notes_row,
        start_column=1,
        end_row=notes_row,
        end_column=6,
    )
    sheet.cell(
        row=notes_row,
        column=1,
        value="Review Notes",
    )
    _style_section_header(
        sheet.cell(
            row=notes_row,
            column=1,
        )
    )

    sheet.merge_cells(
        start_row=notes_row + 1,
        start_column=1,
        end_row=notes_row + 3,
        end_column=6,
    )
    notes_cell = sheet.cell(
        row=notes_row + 1,
        column=1,
        value=(
            summary.review_notes
            or "No special review notes."
        ),
    )
    notes_cell.fill = PatternFill(
        "solid",
        fgColor=NOTE_FILL,
    )
    notes_cell.alignment = Alignment(
        wrap_text=True,
        vertical="top",
    )

    back_cell = sheet.cell(
        row=notes_row + 5,
        column=1,
    )
    _set_internal_hyperlink(
        back_cell,
        location="'Dashboard'!A1",
        display="Back to Dashboard",
    )

    _set_widths(
        sheet,
        {
            "A": 8,
            "B": 34,
            "C": 18,
            "D": 34,
            "E": 15,
            "F": 42,
        },
    )

    sheet.freeze_panes = "A13"


def build_workbook(
    summaries: list[PackageSummary],
    lines: list[PackageLine],
) -> Workbook:
    workbook = Workbook()

    default_sheet = workbook.active
    workbook.remove(default_sheet)

    lines_by_package: dict[str, list[PackageLine]] = (
        defaultdict(list)
    )
    for line in lines:
        lines_by_package[
            line.package_id
        ].append(line)

    _write_dashboard(
        workbook,
        summaries,
    )

    for summary in summaries:
        _write_package_sheet(
            workbook,
            summary,
            lines_by_package.get(
                summary.package_id,
                [],
            ),
        )

    workbook.active = 0
    return workbook


def export_posting_package(
    *,
    summary_path: Path,
    lines_path: Path,
    output_path: Path,
) -> Path:
    summaries = load_summary(summary_path)
    lines = load_lines(lines_path)

    workbook = build_workbook(
        summaries,
        lines,
    )

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )
    workbook.save(output_path)
    return output_path
